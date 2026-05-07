from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def export_to_excel(model: dict) -> bytes:
    wb = Workbook()
    _write_equipment_sheet(wb, model)
    _write_classes_sheet(wb, model)

    if model.get("warnings"):
        _write_warnings_sheet(wb, model)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_warnings_sheet(wb: Workbook, model: dict):
    ws = wb.create_sheet(title="Warnings")
    ws.append(["warning"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for w in model["warnings"]:
        ws.append([w])


def _flatten_equipment(equipment_list) -> list:
    flat = []
    stack = list(equipment_list)
    while stack:
        eq = stack.pop(0)
        flat.append(eq)
        stack = list(eq.children) + stack
    return flat


def _write_equipment_sheet(wb: Workbook, model: dict):
    ws = wb.active
    ws.title = "Equipment"

    flat = _flatten_equipment(model["equipment"])

    # Collect all property names (stable order)
    prop_names = []
    seen = set()
    for eq in flat:
        for p in eq.properties:
            if p.name not in seen:
                prop_names.append(p.name)
                seen.add(p.name)

    # Headers
    headers = ["full_name", "level", "class_ids"]
    for name in prop_names:
        headers.append(name)  # value
        headers.append(f"{name}_uom")  # unit of measure
        headers.append(f"{name}_source")  # provenance

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Rows
    for eq in flat:
        prop_map = {p.name: p for p in eq.properties}

        row = [
            eq.full_name or "",
            eq.level or "",
            ", ".join(eq.class_ids),
        ]

        for name in prop_names:
            prop = prop_map.get(name)
            if prop:
                # value
                row.append("" if prop.value is None else prop.value)

                # uom (normalized if available)
                uom = (
                    prop.normalized_unit_of_measure
                    if getattr(prop, "normalized_unit_of_measure", None)
                    else prop.unit_of_measure
                )
                row.append(uom or "")

                # provenance
                row.append(prop.source or "")
            else:
                # value, uom, source
                row.extend(["", "", ""])

        ws.append(row)


def _write_classes_sheet(wb: Workbook, model: dict):
    ws = wb.create_sheet(title="Classes")

    classes = model["classes"]

    # Collect all class property names
    prop_names = []
    seen = set()
    for cls in classes:
        for p in cls.properties:
            if p.name not in seen:
                prop_names.append(p.name)
                seen.add(p.name)

    # Headers
    headers = ["name", "parent"]
    for name in prop_names:
        headers.append(name)
        headers.append(f"{name}_uom")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Rows
    for cls in classes:
        prop_map = {p.name: p for p in cls.properties}

        row = [cls.name, cls.parent or ""]

        for name in prop_names:
            prop = prop_map.get(name)
            if prop:
                # value
                row.append("" if prop.value is None else prop.value)

                # uom (normalized if available)
                uom = (
                    prop.normalized_unit_of_measure
                    if getattr(prop, "normalized_unit_of_measure", None)
                    else prop.unit_of_measure
                )
                row.append(uom or "")
            else:
                # value, uom
                row.extend(["", ""])

        ws.append(row)
